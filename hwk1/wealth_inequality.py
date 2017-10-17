import numpy as np
import requests
import csv
from StringIO import StringIO
from urllib import urlopen
from zipfile import ZipFile
from collections import defaultdict
from openpyxl import load_workbook
from collections import defaultdict
import matplotlib.pyplot as plt

# pip install openpyxl
#Using the list of countries by continent from World Atlas data, load in the countries.csv file into a pandas DataFrame and name this data set as countries. This data set can be found on Github in the 2014_data repository here.


#Using the data available on Gapminder, load in the Income per person (GDP/capita, PPP$ inflation-adjusted) as a pandas DataFrame and name this data set as income.


#Hint: Consider using the pandas function pandas.read_excel() to read in the .xlsx file directly.
#Read file from path

#Read country CSV
def get_list_of_countries():
    list_of_countries = []
    csv_url = "https://raw.githubusercontent.com/cs109/2014_data/master/countries.csv"
    r = requests.get(csv_url)
    reader = csv.reader(r.iter_lines(), delimiter=',', quotechar='"')
    for row in reader:
        list_of_countries.append(row[0])

    return list_of_countries


def get_list_of_regions():
    list_of_regions = []
    csv_url = "https://raw.githubusercontent.com/cs109/2014_data/master/countries.csv"
    r = requests.get(csv_url)
    reader = csv.reader(r.iter_lines(), delimiter=',', quotechar='"')
    for row in reader:
        list_of_regions.append(row[1])

    return list_of_regions


def get_country_region_map():
    country_region_map = {}
    csv_url = "https://raw.githubusercontent.com/cs109/2014_data/master/countries.csv"
    r = requests.get(csv_url)
    reader = csv.reader(r.iter_lines(), delimiter=',', quotechar='"')
    for row in reader:
        key = row[0]
        value = row[1]
        get_country_region_map[key] = value

    return country_region_map


def get_ipp_lookup():
    #http://zetcode.com/articles/openpyxl/
    income_xlsx_data_path = 'data/income_per_person.xlsx'
    workbook = load_workbook(income_xlsx_data_path)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)

    '''
    for i, row in enumerate(worksheet.iter_rows()):
        for j, cell in enumerate(row):
            if j == 0 and cell.value != None:
                print cell.value
    '''

    #Transform the data set to have years as the rows and countries as the columns. Show the head of this data set when it is loaded.
    #Store data in universally accessible format (hashtable), which can be used to - re-write another xlsx

    country_year_ipp_lookup = defaultdict(int)
    year_index_map = defaultdict(int)
    for i, row in enumerate(worksheet.iter_rows()):
        country_label = None
        for j, cell in enumerate(row):

            # Get year index mapping
            if i == 0 and j > 0:
                year_index_map[j] = int(cell.value)

            # Fill map
            if i > 0:
                #Get Country Value
                if j == 0:
                    country_label = cell.value

                if j > 1 and i > 1 and country_label != None and cell.value != None:
                    year = year_index_map[j]
                    country_year_ipp_lookup[(country_label, year)] = int(cell.value)

    #print year_index_map
    #print country_year_ipp_lookup
    return country_year_ipp_lookup



#Graphically display the distribution of income per person across all countries in the world for any given year (e.g. 2000). What kind of plot would be best?

def graph_income_distribution_per_year(country_year_ipp_lookup, list_of_countries, target_year):
    hist_values = []
    associated_country_labels = []
    country_set = set(list_of_countries)
    for country_year_key in country_year_ipp_lookup:
        country, year = country_year_key
        if year == target_year and country in country_set:
            ipp_value = country_year_ipp_lookup[country_year_key]
            hist_values.append(ipp_value)
            associated_country_labels.append(country)

    '''
    #x = np.random.normal(size = 1000)
    plt.hist(hist_values, normed=True, bins=len(associated_country_labels))
    plt.ylabel('IPP')
    plt.show()
    '''

    #Bar plot to add labels to the x axis, we need sorted stuff and the height
    '''
    x = np.arange(3)
    plt.bar(x, height= [1,2,3])
    plt.xticks(x+.5, ['a','b','c']);
    '''

country_list = get_list_of_countries()
ipp_lookup = get_ipp_lookup()
target_year = 2000 # Be thoughtful about edgecases in future interviews, filter for bad input values

graph_income_distribution_per_year(ipp_lookup, country_list, target_year)



#Write a function to merge the countries and income data sets for any given year.
'''
a DataFrame
   A pandas DataFrame with three columns titled 
   'Country', 'Region', and 'Income'. 
'''


#Create a new feature vector for this purpose

def get_country_region_income_datas(country_list, region_lookup, ipp_lookup, year):
    for country in country_list:
        region = region_lookup[country]
        income = ipp_lookup[(country, year)]
        data_vector = np.array([country, region, income])
    return data_vector


region_lookup = get_country_region_map()
get_country_region_income_datas(country_list, region_lookup, ipp_lookup, 2000)


#Use exploratory data analysis tools such as histograms and boxplots to explore the distribution of the income per person by region data set from 2(c) for a given year. Describe how these change through the recent years?
#Hint: Use a for loop to consider multiple years.


# Graph histogram, Graph Boxplots, to find distribution by region for a year
def graph_ipp_data_by_region_year(country_list, ipp_lookup, year):
    return True








